from django.shortcuts import render, get_object_or_404, redirect
from .models import Item
from .forms import ItemForm

# alteração
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment
import openpyxl
from django.contrib import messages


def home(request):
    return render(request, "home.html")


def listar_itens(request):
    itens = Item.objects.all()
    return render(request, "listar_itens.html", {"itens": itens})


def editar_item(request, id):
    item = get_object_or_404(Item, id=id)

    if request.method == "POST":
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect("listar_itens")
        # Caso o form seja inválido, ele cai aqui e exibe os erros no modal
    else:
        form = ItemForm(instance=item)

    return render(request, "editar_item_modal.html", {"form": form, "item": item})


def adicionar_item(request):
    if request.method == "POST":
        form = ItemForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("listar_itens")
    else:
        form = ItemForm()
    return render(request, "adicionar_item.html", {"form": form})


def excluir_item(request, id):
    item = get_object_or_404(Item, id=id)
    item.delete()
    return redirect("listar_itens")


# alteração
def exportar_xlsx(request):
    # Cria um novo workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Itens"

    # Adiciona o cabeçalho
    ws["A1"] = "Nome"
    ws["B1"] = "Espécime"
    ws["C1"] = "Data de Coleta"

    # Aplica o alinhamento centralizado no cabeçalho
    for cell in ws["1:1"]:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Pega todos os itens cadastrados no banco de dados
    itens = Item.objects.all()

    # Preenche os dados dos itens
    for row_num, item in enumerate(itens, start=2):
        ws[f"A{row_num}"] = item.nome
        ws[f"B{row_num}"] = item.especime
        ws[f"C{row_num}"] = item.data_coleta.strftime(
            "%Y-%m-%d"
        )  # Formata a data para YYYY-MM-DD

    # Configura a resposta HTTP para download do arquivo .xlsx
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="itens_planilha.xlsx"'

    # Salva o arquivo no objeto HttpResponse
    wb.save(response)

    return response


def importar_xlsx(request):
    if request.method == "POST" and request.FILES.get("arquivo_excel"):
        arquivo = request.FILES["arquivo_excel"]
        wb = openpyxl.load_workbook(arquivo)
        sheet = wb.active

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            print(f"Linha {i}: {row}")  # Imprime o conteúdo da linha para debug

            # Garante que apenas as 3 primeiras colunas sejam usadas
            nome, especime, data_coleta = row[:3]

            # Verifica se todos os valores são válidos
            if nome and especime and data_coleta:
                Item.objects.create(
                    nome=nome, especime=especime, data_coleta=data_coleta
                )
            else:
                messages.warning(request, f"Linha {i} ignorada por dados incompletos.")

        messages.success(request, "Importação concluída com sucesso!")
        return redirect("listar_itens")

    return render(request, "importar_xlsx.html")


def excluir_itens_massa(request):
    if request.method == "POST":
        # Recebe os itens selecionados
        itens_selecionados = request.POST.getlist("itens_selecionados")

        if itens_selecionados:
            # Converte os IDs em inteiros e exclui os itens
            itens = Item.objects.filter(id__in=itens_selecionados)
            quantidade_deletada = itens.delete()[
                0
            ]  # .delete() retorna tupla (quantidade_deletada, {model: deleted_count})

            if quantidade_deletada > 0:
                messages.success(
                    request, f"{quantidade_deletada} itens excluídos com sucesso."
                )
            else:
                messages.warning(request, "Nenhum item foi selecionado para exclusão.")
        else:
            messages.warning(request, "Nenhum item foi selecionado para exclusão.")

        return redirect("listar_itens")
